from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

app = Flask(__name__)
app.config['SECRET_KEY'] = 'tu-clave-secreta-aqui'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///sistema_ventas.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Modelo de Usuario
class Usuario(UserMixin, db.Model):
    __tablename__ = 'usuario'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    ventas = db.relationship('Venta', backref='vendedor')

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

# Modelo de Categoría
class Categoria(db.Model):
    __tablename__ = 'categoria'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    descripcion = db.Column(db.Text)
    productos = db.relationship('Producto', backref='categoria', lazy=True)

# Modelo de Producto
class Producto(db.Model):
    __tablename__ = 'producto'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    descripcion = db.Column(db.Text)
    precio = db.Column(db.Float, nullable=False)  # Precio de venta
    precio_compra = db.Column(db.Float, nullable=False)  # Precio de compra
    categoria_id = db.Column(db.Integer, db.ForeignKey('categoria.id'), nullable=False)
    stock = db.relationship('Stock', backref='producto', uselist=False)
    ventas = db.relationship('Venta', secondary='venta_producto', back_populates='productos')
    
    def margen_ganancia(self):
        """Calcula el margen de ganancia en porcentaje"""
        if self.precio_compra > 0:
            return ((self.precio - self.precio_compra) / self.precio_compra) * 100
        return 0
    
    def ganancia_unitaria(self):
        """Calcula la ganancia por unidad"""
        return self.precio - self.precio_compra

# Modelo de Stock
class Stock(db.Model):
    __tablename__ = 'stock'
    id = db.Column(db.Integer, primary_key=True)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    cantidad_disponible = db.Column(db.Integer, default=0)
    cantidad_minima = db.Column(db.Integer, default=5)
    fecha_actualizacion = db.Column(db.DateTime, default=datetime.utcnow)

# Modelo de Cliente
class Cliente(db.Model):
    __tablename__ = 'cliente'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    telefono = db.Column(db.String(20))
    ventas = db.relationship('Venta', backref='cliente')

# Modelo de Lugar de Entrega
class LugarEntrega(db.Model):
    __tablename__ = 'lugar_entrega'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    direccion = db.Column(db.Text)
    telefono = db.Column(db.String(20))
    tipo = db.Column(db.String(50), default='domicilio')  # domicilio, oficina, tienda, etc.
    ventas = db.relationship('Venta', backref='lugar_entrega')

# Modelo de Descuento
class Descuento(db.Model):
    __tablename__ = 'descuento'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    porcentaje = db.Column(db.Float, nullable=False)
    fecha_inicio = db.Column(db.DateTime, nullable=False)
    fecha_fin = db.Column(db.DateTime, nullable=False)
    activo = db.Column(db.Boolean, default=True)
    ventas = db.relationship('Venta', backref='descuento')

# Tabla de relación muchos a muchos entre Venta y Producto
venta_producto = db.Table('venta_producto',
    db.Column('venta_id', db.Integer, db.ForeignKey('venta.id'), primary_key=True),
    db.Column('producto_id', db.Integer, db.ForeignKey('producto.id'), primary_key=True),
    db.Column('cantidad', db.Integer, nullable=False),
    db.Column('precio_unitario', db.Float, nullable=False)
)

# Modelo de Venta (entidad central)
class Venta(db.Model):
    __tablename__ = 'venta'
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    total = db.Column(db.Float, nullable=False)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    lugar_entrega_id = db.Column(db.Integer, db.ForeignKey('lugar_entrega.id'), nullable=False)
    vendedor_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    estado = db.Column(db.String(20), default='contraentrega')  # contraentrega, cancelado, abonado
    descuento_id = db.Column(db.Integer, db.ForeignKey('descuento.id'))
    productos = db.relationship('Producto', secondary='venta_producto', back_populates='ventas')
    
    def ganancia_total(self):
        """Calcula la ganancia total de la venta"""
        ganancia = 0
        for producto in self.productos:
            # Obtener la cantidad vendida de este producto
            cantidad = db.session.query(venta_producto.c.cantidad).filter_by(
                venta_id=self.id, producto_id=producto.id
            ).scalar()
            if cantidad:
                ganancia += producto.ganancia_unitaria() * cantidad
        return ganancia

# Modelo de Ganancias (para tracking histórico)
class Ganancias(db.Model):
    __tablename__ = 'ganancias'
    id = db.Column(db.Integer, primary_key=True)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    venta_id = db.Column(db.Integer, db.ForeignKey('venta.id'), nullable=False)
    cantidad_vendida = db.Column(db.Integer, nullable=False)
    precio_venta = db.Column(db.Float, nullable=False)
    precio_compra = db.Column(db.Float, nullable=False)
    ganancia_unitaria = db.Column(db.Float, nullable=False)
    ganancia_total = db.Column(db.Float, nullable=False)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    
    producto = db.relationship('Producto', backref='ganancias')
    venta = db.relationship('Venta', backref='ganancias')

@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

# Rutas
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = Usuario.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user)
            flash('¡Inicio de sesión exitoso!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        
        # Verificar si el usuario ya existe
        if Usuario.query.filter_by(username=username).first():
            flash('El nombre de usuario ya existe', 'error')
            return render_template('register.html')
        
        if Usuario.query.filter_by(email=email).first():
            flash('El email ya está registrado', 'error')
            return render_template('register.html')
        
        # Crear nuevo usuario
        user = Usuario(username=username, email=email)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        
        flash('¡Registro exitoso! Ahora puedes iniciar sesión', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/dashboard')
@login_required
def dashboard():
    # Estadísticas básicas
    total_productos = Producto.query.count()
    total_clientes = Cliente.query.count()
    total_ventas = Venta.query.count()
    total_lugares = LugarEntrega.query.count()
    
    return render_template('dashboard.html', 
                         user=current_user,
                         total_productos=total_productos,
                         total_clientes=total_clientes,
                         total_ventas=total_ventas,
                         total_lugares=total_lugares)

# Rutas para Productos
@app.route('/productos')
@login_required
def productos():
    productos = Producto.query.join(Categoria).all()
    return render_template('productos.html', productos=productos)

@app.route('/productos/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_producto():
    if request.method == 'POST':
        nombre = request.form['nombre']
        descripcion = request.form['descripcion']
        precio = float(request.form['precio'])
        precio_compra = float(request.form['precio_compra'])
        categoria_id = int(request.form['categoria_id'])
        cantidad_stock = int(request.form['cantidad_stock'])
        
        # Crear producto
        producto = Producto(nombre=nombre, descripcion=descripcion, 
                           precio=precio, precio_compra=precio_compra, categoria_id=categoria_id)
        db.session.add(producto)
        db.session.flush()  # Para obtener el ID del producto
        
        # Crear stock
        stock = Stock(producto_id=producto.id, cantidad_disponible=cantidad_stock)
        db.session.add(stock)
        db.session.commit()
        
        flash('Producto creado exitosamente', 'success')
        return redirect(url_for('productos'))
    
    categorias = Categoria.query.all()
    return render_template('nuevo_producto.html', categorias=categorias)

@app.route('/productos/editar/<int:producto_id>', methods=['GET', 'POST'])
@login_required
def editar_producto(producto_id):
    producto = Producto.query.get_or_404(producto_id)
    
    if request.method == 'POST':
        producto.nombre = request.form['nombre']
        producto.descripcion = request.form['descripcion']
        producto.precio = float(request.form['precio'])
        producto.precio_compra = float(request.form['precio_compra'])
        producto.categoria_id = int(request.form['categoria_id'])
        
        db.session.commit()
        flash('Producto actualizado exitosamente', 'success')
        return redirect(url_for('productos'))
    
    categorias = Categoria.query.all()
    return render_template('editar_producto.html', producto=producto, categorias=categorias)

@app.route('/productos/eliminar/<int:producto_id>')
@login_required
def eliminar_producto(producto_id):
    producto = Producto.query.get_or_404(producto_id)
    
    # Eliminar stock asociado
    if producto.stock:
        db.session.delete(producto.stock)
    
    # Eliminar ganancias asociadas
    Ganancias.query.filter_by(producto_id=producto_id).delete()
    
    # Eliminar de ventas (tabla intermedia)
    db.session.execute(venta_producto.delete().where(venta_producto.c.producto_id == producto_id))
    
    db.session.delete(producto)
    db.session.commit()
    
    flash('Producto eliminado exitosamente', 'success')
    return redirect(url_for('productos'))

# Rutas para Categorías
@app.route('/categorias')
@login_required
def categorias():
    categorias = Categoria.query.all()
    return render_template('categorias.html', categorias=categorias)

@app.route('/categorias/nueva', methods=['GET', 'POST'])
@login_required
def nueva_categoria():
    if request.method == 'POST':
        nombre = request.form['nombre']
        descripcion = request.form['descripcion']
        
        categoria = Categoria(nombre=nombre, descripcion=descripcion)
        db.session.add(categoria)
        db.session.commit()
        
        flash('Categoría creada exitosamente', 'success')
        return redirect(url_for('categorias'))
    
    return render_template('nueva_categoria.html')

@app.route('/categorias/editar/<int:categoria_id>', methods=['GET', 'POST'])
@login_required
def editar_categoria(categoria_id):
    categoria = Categoria.query.get_or_404(categoria_id)
    
    if request.method == 'POST':
        categoria.nombre = request.form['nombre']
        categoria.descripcion = request.form['descripcion']
        
        db.session.commit()
        flash('Categoría actualizada exitosamente', 'success')
        return redirect(url_for('categorias'))
    
    return render_template('editar_categoria.html', categoria=categoria)

@app.route('/categorias/eliminar/<int:categoria_id>')
@login_required
def eliminar_categoria(categoria_id):
    categoria = Categoria.query.get_or_404(categoria_id)
    
    # Verificar si hay productos asociados
    if categoria.productos:
        flash('No se puede eliminar la categoría porque tiene productos asociados', 'error')
        return redirect(url_for('categorias'))
    
    db.session.delete(categoria)
    db.session.commit()
    
    flash('Categoría eliminada exitosamente', 'success')
    return redirect(url_for('categorias'))

# Rutas para Clientes
@app.route('/clientes')
@login_required
def clientes():
    clientes = Cliente.query.all()
    return render_template('clientes.html', clientes=clientes)

@app.route('/clientes/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_cliente():
    if request.method == 'POST':
        try:
            nombre = request.form['nombre']
            telefono = request.form['telefono']
            
            cliente = Cliente(nombre=nombre, telefono=telefono)
            db.session.add(cliente)
            db.session.commit()
            
            flash('Cliente creado exitosamente', 'success')
            return redirect(url_for('clientes'))
        except Exception as e:
            db.session.rollback()
            flash('Error al crear el cliente: ' + str(e), 'error')
            return render_template('nuevo_cliente.html')
    
    return render_template('nuevo_cliente.html')

@app.route('/clientes/editar/<int:cliente_id>', methods=['GET', 'POST'])
@login_required
def editar_cliente(cliente_id):
    cliente = Cliente.query.get_or_404(cliente_id)
    
    if request.method == 'POST':
        try:
            cliente.nombre = request.form['nombre']
            cliente.telefono = request.form['telefono']
            
            db.session.commit()
            flash('Cliente actualizado exitosamente', 'success')
            return redirect(url_for('clientes'))
        except Exception as e:
            db.session.rollback()
            flash('Error al actualizar el cliente: ' + str(e), 'error')
            return render_template('editar_cliente.html', cliente=cliente)
    
    return render_template('editar_cliente.html', cliente=cliente)

@app.route('/clientes/eliminar/<int:cliente_id>')
@login_required
def eliminar_cliente(cliente_id):
    cliente = Cliente.query.get_or_404(cliente_id)
    
    # Verificar si hay ventas asociadas
    if cliente.ventas:
        flash('No se puede eliminar el cliente porque tiene ventas asociadas', 'error')
        return redirect(url_for('clientes'))
    
    db.session.delete(cliente)
    db.session.commit()
    
    flash('Cliente eliminado exitosamente', 'success')
    return redirect(url_for('clientes'))

# Rutas para Lugares de Entrega
@app.route('/lugares-entrega')
@login_required
def lugares_entrega():
    lugares = LugarEntrega.query.all()
    return render_template('lugares_entrega.html', lugares=lugares)

@app.route('/lugares-entrega/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_lugar_entrega():
    if request.method == 'POST':
        nombre = request.form['nombre']
        direccion = request.form['direccion']
        telefono = request.form['telefono']
        tipo = request.form['tipo']
        
        lugar = LugarEntrega(nombre=nombre, direccion=direccion, telefono=telefono, tipo=tipo)
        db.session.add(lugar)
        db.session.commit()
        
        flash('Lugar de entrega creado exitosamente', 'success')
        return redirect(url_for('lugares_entrega'))
    
    return render_template('nuevo_lugar_entrega.html')

@app.route('/lugares-entrega/editar/<int:lugar_id>', methods=['GET', 'POST'])
@login_required
def editar_lugar_entrega(lugar_id):
    lugar = LugarEntrega.query.get_or_404(lugar_id)
    
    if request.method == 'POST':
        lugar.nombre = request.form['nombre']
        lugar.direccion = request.form['direccion']
        lugar.telefono = request.form['telefono']
        lugar.tipo = request.form['tipo']
        
        db.session.commit()
        flash('Lugar de entrega actualizado exitosamente', 'success')
        return redirect(url_for('lugares_entrega'))
    
    return render_template('editar_lugar_entrega.html', lugar=lugar)

@app.route('/lugares-entrega/eliminar/<int:lugar_id>')
@login_required
def eliminar_lugar_entrega(lugar_id):
    lugar = LugarEntrega.query.get_or_404(lugar_id)
    
    # Verificar si hay ventas asociadas
    if lugar.ventas:
        flash('No se puede eliminar el lugar porque tiene ventas asociadas', 'error')
        return redirect(url_for('lugares_entrega'))
    
    db.session.delete(lugar)
    db.session.commit()
    
    flash('Lugar de entrega eliminado exitosamente', 'success')
    return redirect(url_for('lugares_entrega'))

# Rutas para Ventas
@app.route('/ventas')
@login_required
def ventas():
    ventas = Venta.query.join(Cliente).join(LugarEntrega).join(Usuario).all()
    return render_template('ventas.html', ventas=ventas)

@app.route('/ventas/nueva', methods=['GET', 'POST'])
@login_required
def nueva_venta():
    # Validar que solo se puedan crear ventas en el día actual
    from datetime import date
    hoy = date.today()
    
    if request.method == 'POST':
        cliente_id = int(request.form['cliente_id'])
        lugar_entrega_id = int(request.form['lugar_entrega_id'])
        vendedor_id = int(request.form['vendedor_id'])
        estado = request.form['estado']
        descuento_id = request.form.get('descuento_id') or None
        
        # Crear venta
        venta = Venta(cliente_id=cliente_id, lugar_entrega_id=lugar_entrega_id, 
                     vendedor_id=vendedor_id, estado=estado, descuento_id=descuento_id, total=0)
        db.session.add(venta)
        db.session.flush()
        
        # Procesar productos
        total = 0
        for key, value in request.form.items():
            if key.startswith('producto_') and value:
                producto_id = int(key.split('_')[1])
                cantidad = int(value)
                producto = Producto.query.get(producto_id)
                
                if producto and producto.stock.cantidad_disponible >= cantidad:
                    # Agregar producto a la venta
                    db.session.execute(venta_producto.insert().values(
                        venta_id=venta.id,
                        producto_id=producto_id,
                        cantidad=cantidad,
                        precio_unitario=producto.precio
                    ))
                    
                    # Registrar ganancia
                    ganancia_unitaria = producto.ganancia_unitaria()
                    ganancia_total = ganancia_unitaria * cantidad
                    
                    ganancia = Ganancias(
                        producto_id=producto_id,
                        venta_id=venta.id,
                        cantidad_vendida=cantidad,
                        precio_venta=producto.precio,
                        precio_compra=producto.precio_compra,
                        ganancia_unitaria=ganancia_unitaria,
                        ganancia_total=ganancia_total
                    )
                    db.session.add(ganancia)
                    
                    # Actualizar stock
                    producto.stock.cantidad_disponible -= cantidad
                    total += producto.precio * cantidad
        
        # Aplicar descuento si existe
        if descuento_id:
            descuento = Descuento.query.get(descuento_id)
            if descuento and descuento.activo:
                total = total * (1 - descuento.porcentaje / 100)
        
        venta.total = total
        db.session.commit()
        
        flash('Venta realizada exitosamente', 'success')
        return redirect(url_for('ventas'))
    
    clientes = Cliente.query.all()
    lugares_entrega = LugarEntrega.query.all()
    vendedores = Usuario.query.all()
    productos = Producto.query.join(Stock).filter(Stock.cantidad_disponible > 0).all()
    descuentos = Descuento.query.filter_by(activo=True).all()
    
    return render_template('nueva_venta.html', 
                         clientes=clientes, lugares_entrega=lugares_entrega, 
                         vendedores=vendedores, productos=productos, descuentos=descuentos)

# Ruta para exportar ventas a Excel
@app.route('/ventas/exportar')
@login_required
def exportar_ventas():
    ventas = Venta.query.join(Cliente).join(LugarEntrega).join(Usuario).all()
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"
    
    # Encabezados
    headers = ['ID Venta', 'Fecha', 'Cliente', 'Lugar de Entrega', 'Vendedor', 'Estado', 'Total', 'Ganancia Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    for row, venta in enumerate(ventas, 2):
        ws.cell(row=row, column=1, value=venta.id)
        ws.cell(row=row, column=2, value=venta.fecha.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=3, value=venta.cliente.nombre)
        ws.cell(row=row, column=4, value=venta.lugar_entrega.nombre)
        ws.cell(row=row, column=5, value=venta.vendedor.username)
        ws.cell(row=row, column=6, value=venta.estado.title())
        ws.cell(row=row, column=7, value=venta.total)
        ws.cell(row=row, column=8, value=venta.ganancia_total())
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta
    response = make_response()
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_ventas.xlsx'
    
    # Guardar en memoria
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.data = output.getvalue()
    
    return response

# Ruta para importar ventas desde Excel
@app.route('/ventas/importar', methods=['GET', 'POST'])
@login_required
def importar_ventas():
    if request.method == 'POST':
        if 'archivo_excel' not in request.files:
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('ventas'))
        
        archivo = request.files['archivo_excel']
        if archivo.filename == '':
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('ventas'))
        
        if archivo and archivo.filename.endswith(('.xlsx', '.xls')):
            try:
                # Cargar el archivo Excel
                wb = load_workbook(archivo)
                ws = wb.active
                
                ventas_importadas = 0
                errores = []
                
                # Leer datos del Excel (asumiendo formato: Fecha, Cliente, Producto, Cantidad, Precio, Vendedor)
                for row in range(2, ws.max_row + 1):  # Saltar encabezados
                    try:
                        fecha_str = ws.cell(row=row, column=1).value
                        cliente_nombre = ws.cell(row=row, column=2).value
                        producto_nombre = ws.cell(row=row, column=3).value
                        cantidad = ws.cell(row=row, column=4).value
                        precio = ws.cell(row=row, column=5).value
                        vendedor_nombre = ws.cell(row=row, column=6).value
                        
                        if not all([fecha_str, cliente_nombre, producto_nombre, cantidad, precio, vendedor_nombre]):
                            continue
                        
                        # Buscar o crear cliente
                        cliente = Cliente.query.filter_by(nombre=cliente_nombre).first()
                        if not cliente:
                            cliente = Cliente(nombre=cliente_nombre, telefono='')
                            db.session.add(cliente)
                            db.session.flush()
                        
                        # Buscar producto
                        producto = Producto.query.filter_by(nombre=producto_nombre).first()
                        if not producto:
                            errores.append(f"Producto '{producto_nombre}' no encontrado en fila {row}")
                            continue
                        
                        # Verificar stock
                        if producto.stock.cantidad_disponible < cantidad:
                            errores.append(f"Stock insuficiente para '{producto_nombre}' en fila {row}")
                            continue
                        
                        # Buscar vendedor
                        vendedor = Usuario.query.filter_by(username=vendedor_nombre).first()
                        if not vendedor:
                            errores.append(f"Vendedor '{vendedor_nombre}' no encontrado en fila {row}")
                            continue
                        
                        # Crear lugar de entrega por defecto
                        lugar_entrega = LugarEntrega.query.filter_by(nombre='Importado').first()
                        if not lugar_entrega:
                            lugar_entrega = LugarEntrega(
                                nombre='Importado',
                                direccion='Importado desde Excel',
                                telefono='',
                                tipo='importado'
                            )
                            db.session.add(lugar_entrega)
                            db.session.flush()
                        
                        # Crear venta
                        venta = Venta(
                            fecha=datetime.strptime(fecha_str, '%Y-%m-%d') if isinstance(fecha_str, str) else fecha_str,
                            cliente_id=cliente.id,
                            lugar_entrega_id=lugar_entrega.id,
                            vendedor_id=vendedor.id,
                            estado='contraentrega',
                            total=precio * cantidad
                        )
                        db.session.add(venta)
                        db.session.flush()
                        
                        # Agregar producto a la venta
                        db.session.execute(venta_producto.insert().values(
                            venta_id=venta.id,
                            producto_id=producto.id,
                            cantidad=cantidad,
                            precio_unitario=precio
                        ))
                        
                        # Registrar ganancia
                        ganancia_unitaria = precio - producto.precio_compra
                        ganancia_total = ganancia_unitaria * cantidad
                        
                        ganancia = Ganancias(
                            producto_id=producto.id,
                            venta_id=venta.id,
                            cantidad_vendida=cantidad,
                            precio_venta=precio,
                            precio_compra=producto.precio_compra,
                            ganancia_unitaria=ganancia_unitaria,
                            ganancia_total=ganancia_total
                        )
                        db.session.add(ganancia)
                        
                        # Actualizar stock
                        producto.stock.cantidad_disponible -= cantidad
                        
                        ventas_importadas += 1
                        
                    except Exception as e:
                        errores.append(f"Error en fila {row}: {str(e)}")
                        continue
                
                db.session.commit()
                
                if ventas_importadas > 0:
                    flash(f'Se importaron {ventas_importadas} ventas exitosamente', 'success')
                if errores:
                    flash(f'Errores encontrados: {len(errores)}', 'warning')
                    for error in errores[:5]:  # Mostrar solo los primeros 5 errores
                        flash(error, 'error')
                
                return redirect(url_for('ventas'))
                
            except Exception as e:
                flash(f'Error al procesar el archivo: {str(e)}', 'error')
                return redirect(url_for('ventas'))
        else:
            flash('Formato de archivo no válido. Use archivos .xlsx o .xls', 'error')
            return redirect(url_for('ventas'))
    
    # Obtener productos y vendedores para mostrar en la plantilla
    productos = Producto.query.all()
    vendedores = Usuario.query.all()
    return render_template('importar_ventas.html', productos=productos, vendedores=vendedores)

# Ruta para descargar plantilla de Excel
@app.route('/ventas/plantilla')
@login_required
def descargar_plantilla():
    # Crear libro de Excel con plantilla
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Ventas"
    
    # Encabezados
    headers = ['Fecha', 'Cliente', 'Producto', 'Cantidad', 'Precio', 'Vendedor']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Datos de ejemplo
    ejemplos = [
        ['2024-01-15', 'Juan Pérez', 'Producto A', 2, 25.50, 'Alonso'],
        ['2024-01-16', 'María García', 'Producto B', 1, 15.00, 'Andrea'],
        ['2024-01-17', 'Carlos López', 'Producto A', 3, 25.50, 'Alonso']
    ]
    
    for row, ejemplo in enumerate(ejemplos, 2):
        for col, valor in enumerate(ejemplo, 1):
            ws.cell(row=row, column=col, value=valor)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta
    response = make_response()
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=plantilla_ventas.xlsx'
    
    # Guardar en memoria
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.data = output.getvalue()
    
    return response

# Rutas para Ganancias
@app.route('/ganancias')
@login_required
def ganancias():
    # Estadísticas generales
    total_ganancias = db.session.query(db.func.sum(Ganancias.ganancia_total)).scalar() or 0
    total_ventas = Venta.query.count()
    ganancia_promedio = total_ganancias / total_ventas if total_ventas > 0 else 0
    
    # Todas las ventas individuales con información del vendedor (excluyendo cantidad 0)
    ventas_detalladas = db.session.query(
        Ganancias.id,
        Ganancias.fecha,
        Producto.nombre.label('producto_nombre'),
        Cliente.nombre.label('cliente_nombre'),
        Usuario.username.label('vendedor_nombre'),
        Ganancias.cantidad_vendida,
        Ganancias.precio_venta,
        Ganancias.precio_compra,
        Ganancias.ganancia_unitaria,
        Ganancias.ganancia_total,
        Venta.total.label('venta_total'),
        Venta.estado
    ).join(Producto).join(Venta).join(Cliente).join(Usuario).filter(
        Ganancias.cantidad_vendida > 0
    ).order_by(Ganancias.fecha.desc()).all()
    
    # Ganancias por producto (sin duplicados) - para estadísticas (excluyendo cantidad 0)
    ganancias_por_producto_raw = db.session.query(
        Producto.id,
        Producto.nombre,
        Producto.precio,
        Producto.precio_compra,
        db.func.sum(Ganancias.ganancia_total).label('ganancia_total'),
        db.func.sum(Ganancias.cantidad_vendida).label('cantidad_vendida'),
        db.func.avg(Ganancias.ganancia_unitaria).label('ganancia_promedio'),
        (Producto.precio - Producto.precio_compra).label('diferencia_precio')
    ).join(Ganancias).filter(
        Ganancias.cantidad_vendida > 0
    ).group_by(Producto.id, Producto.nombre, Producto.precio, Producto.precio_compra).all()
    
    # Convertir a diccionarios para JSON
    ganancias_por_producto = []
    for row in ganancias_por_producto_raw:
        ganancias_por_producto.append({
            'id': row.id,
            'nombre': row.nombre,
            'precio': float(row.precio),
            'precio_compra': float(row.precio_compra),
            'ganancia_total': float(row.ganancia_total or 0),
            'cantidad_vendida': int(row.cantidad_vendida or 0),
            'ganancia_promedio': float(row.ganancia_promedio or 0),
            'diferencia_precio': float(row.diferencia_precio or 0)
        })
    
    # Calcular ganancias diarias para la gráfica lineal
    ganancias_diarias_raw = db.session.query(
        db.func.date(db.func.datetime(Ganancias.fecha, '-5 hours')).label('fecha'),
        db.func.sum(Ganancias.ganancia_total).label('ganancia_diaria')
    ).filter(
        Ganancias.cantidad_vendida > 0
    ).group_by(db.func.date(db.func.datetime(Ganancias.fecha, '-5 hours'))).order_by(
        db.func.date(db.func.datetime(Ganancias.fecha, '-5 hours'))
    ).all()
    
    # Convertir a diccionarios para JSON
    ganancias_diarias = []
    for row in ganancias_diarias_raw:
        ganancias_diarias.append({
            'fecha': str(row.fecha),
            'ganancia_diaria': float(row.ganancia_diaria or 0)
        })
    
    # Consulta para ganancias en tiempo real del día actual (por venta individual)
    from datetime import datetime, date
    hoy = date.today()
    ganancias_tiempo_real_raw = db.session.query(
        Ganancias.fecha.label('fecha_venta'),
        Ganancias.ganancia_total.label('ganancia_venta'),
        Producto.nombre.label('producto_nombre'),
        Cliente.nombre.label('cliente_nombre'),
        Usuario.username.label('vendedor_nombre')
    ).join(Producto).join(Venta).join(Cliente).join(Usuario).filter(
        db.func.date(Ganancias.fecha) == hoy,
        Ganancias.cantidad_vendida > 0
    ).order_by(Ganancias.fecha).all()
    
    # Convertir ganancias en tiempo real a diccionarios
    ganancias_tiempo_real = []
    ganancia_acumulada = 0
    for row in ganancias_tiempo_real_raw:
        ganancia_acumulada += float(row.ganancia_venta or 0)
        ganancias_tiempo_real.append({
            'fecha_venta': row.fecha_venta.isoformat() if row.fecha_venta else None,
            'ganancia_venta': float(row.ganancia_venta or 0),
            'ganancia_acumulada': ganancia_acumulada,
            'producto_nombre': row.producto_nombre,
            'cliente_nombre': row.cliente_nombre,
            'vendedor_nombre': row.vendedor_nombre
        })
    
    # Convertir ventas detalladas a diccionarios
    ventas_detalladas_dict = []
    for row in ventas_detalladas:
        ventas_detalladas_dict.append({
            'id': row.id,
            'fecha': row.fecha.isoformat() if row.fecha else None,
            'producto_nombre': row.producto_nombre,
            'cliente_nombre': row.cliente_nombre,
            'vendedor_nombre': row.vendedor_nombre,
            'cantidad_vendida': int(row.cantidad_vendida),
            'precio_venta': float(row.precio_venta),
            'precio_compra': float(row.precio_compra),
            'ganancia_unitaria': float(row.ganancia_unitaria),
            'ganancia_total': float(row.ganancia_total),
            'venta_total': float(row.venta_total or 0),
            'estado': row.estado
        })
    
    # Calcular total del día actual
    total_hoy = sum(ganancia['ganancia_venta'] for ganancia in ganancias_tiempo_real)
    
    return render_template('ganancias.html', 
                         total_ganancias=total_ganancias,
                         total_ventas=total_ventas,
                         ganancia_promedio=ganancia_promedio,
                         ganancias_por_producto=ganancias_por_producto,
                         ventas_detalladas=ventas_detalladas_dict,
                         ganancias_diarias=ganancias_diarias,
                         ganancias_tiempo_real=ganancias_tiempo_real,
                         total_hoy=total_hoy)

@app.route('/ganancias/data')
@login_required
def ganancias_data():
    """Datos JSON en tiempo real para actualizar las gráficas de ganancias"""
    # Estadísticas generales
    total_ganancias = db.session.query(db.func.sum(Ganancias.ganancia_total)).scalar() or 0
    total_ventas = Venta.query.count()
    ganancia_promedio = total_ganancias / total_ventas if total_ventas > 0 else 0

    # Ganancias por producto (excluyendo cantidad 0)
    ganancias_por_producto_raw = db.session.query(
        Producto.id,
        Producto.nombre,
        Producto.precio,
        Producto.precio_compra,
        db.func.sum(Ganancias.ganancia_total).label('ganancia_total'),
        db.func.sum(Ganancias.cantidad_vendida).label('cantidad_vendida'),
        db.func.avg(Ganancias.ganancia_unitaria).label('ganancia_promedio'),
        (Producto.precio - Producto.precio_compra).label('diferencia_precio')
    ).join(Ganancias).filter(
        Ganancias.cantidad_vendida > 0
    ).group_by(Producto.id, Producto.nombre, Producto.precio, Producto.precio_compra).all()

    ganancias_por_producto = []
    for row in ganancias_por_producto_raw:
        ganancias_por_producto.append({
            'id': row.id,
            'nombre': row.nombre,
            'precio': float(row.precio),
            'precio_compra': float(row.precio_compra),
            'ganancia_total': float(row.ganancia_total or 0),
            'cantidad_vendida': int(row.cantidad_vendida or 0),
            'ganancia_promedio': float(row.ganancia_promedio or 0),
            'diferencia_precio': float(row.diferencia_precio or 0)
        })

    # Ganancias diarias (histórico por día) en hora de Perú (UTC-5)
    ganancias_diarias_raw = db.session.query(
        db.func.date(db.func.datetime(Ganancias.fecha, '-5 hours')).label('fecha'),
        db.func.sum(Ganancias.ganancia_total).label('ganancia_diaria')
    ).filter(
        Ganancias.cantidad_vendida > 0
    ).group_by(db.func.date(db.func.datetime(Ganancias.fecha, '-5 hours'))).order_by(
        db.func.date(db.func.datetime(Ganancias.fecha, '-5 hours'))
    ).all()

    ganancias_diarias = []
    for row in ganancias_diarias_raw:
        ganancias_diarias.append({
            'fecha': str(row.fecha),
            'ganancia_diaria': float(row.ganancia_diaria or 0)
        })

    # Ganancias en tiempo real (por venta de hoy)
    from datetime import date
    hoy = date.today()
    ganancias_tiempo_real_raw = db.session.query(
        Ganancias.fecha.label('fecha_venta'),
        Ganancias.ganancia_total.label('ganancia_venta')
    ).filter(
        db.func.date(Ganancias.fecha) == hoy,
        Ganancias.cantidad_vendida > 0
    ).order_by(Ganancias.fecha).all()

    ganancias_tiempo_real = []
    ganancia_acumulada = 0
    for row in ganancias_tiempo_real_raw:
        ganancia_venta = float(row.ganancia_venta or 0)
        ganancia_acumulada += ganancia_venta
        ganancias_tiempo_real.append({
            'fecha_venta': row.fecha_venta.isoformat() if row.fecha_venta else None,
            'ganancia_venta': ganancia_venta,
            'ganancia_acumulada': ganancia_acumulada
        })

    total_hoy = ganancia_acumulada
    ventas_hoy = len(ganancias_tiempo_real)

    return jsonify({
        'total_ganancias': float(total_ganancias),
        'total_ventas': int(total_ventas),
        'ganancia_promedio': float(ganancia_promedio),
        'ganancias_por_producto': ganancias_por_producto,
        'ganancias_diarias': ganancias_diarias,
        'ganancias_tiempo_real': ganancias_tiempo_real,
        'total_hoy': float(total_hoy),
        'ventas_hoy': int(ventas_hoy),
        'currency_symbol': 'S/.'
    })
@app.route('/ganancias/producto/<int:producto_id>')
@login_required
def ganancias_producto(producto_id):
    producto = Producto.query.get_or_404(producto_id)
    ganancias = Ganancias.query.filter_by(producto_id=producto_id).filter(
        Ganancias.cantidad_vendida > 0
    ).order_by(Ganancias.fecha.desc()).all()
    
    # Estadísticas del producto
    total_ganancia = sum(g.ganancia_total for g in ganancias)
    total_vendido = sum(g.cantidad_vendida for g in ganancias)
    margen_promedio = producto.margen_ganancia()
    
    return render_template('ganancias_producto.html',
                         producto=producto,
                         ganancias=ganancias,
                         total_ganancia=total_ganancia,
                         total_vendido=total_vendido,
                         margen_promedio=margen_promedio)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Has cerrado sesión correctamente', 'info')
    return redirect(url_for('login'))

# Rutas para gestión de respaldos
@app.route('/respaldos')
@login_required
def respaldos():
    """Mostrar lista de respaldos disponibles"""
    try:
        import os
        from datetime import datetime
        
        backup_dir = 'backups'
        respaldos = []
        
        if os.path.exists(backup_dir):
            for filename in os.listdir(backup_dir):
                if filename.endswith('.db'):
                    file_path = os.path.join(backup_dir, filename)
                    file_stats = os.stat(file_path)
                    file_size = file_stats.st_size
                    file_date = datetime.fromtimestamp(file_stats.st_mtime)
                    
                    respaldos.append({
                        'nombre': filename,
                        'ruta': file_path,
                        'tamaño': file_size,
                        'fecha': file_date,
                        'fecha_formateada': file_date.strftime('%d/%m/%Y %H:%M:%S')
                    })
            
            # Ordenar por fecha (más recientes primero)
            respaldos.sort(key=lambda x: x['fecha'], reverse=True)
        
        return render_template('respaldos.html', respaldos=respaldos)
    except Exception as e:
        flash(f'Error al cargar respaldos: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/respaldos/crear')
@login_required
def crear_respaldo_manual():
    """Crear respaldo manual"""
    try:
        backup_path = crear_respaldo_automatico()
        if backup_path:
            flash('Respaldo creado exitosamente', 'success')
        else:
            flash('Error al crear el respaldo', 'error')
    except Exception as e:
        flash(f'Error al crear respaldo: {str(e)}', 'error')
    
    return redirect(url_for('respaldos'))

@app.route('/respaldos/restaurar/<filename>')
@login_required
def restaurar_respaldo(filename):
    """Restaurar desde un respaldo"""
    try:
        import shutil
        from flask import request
        
        backup_path = os.path.join('backups', filename)
        if not os.path.exists(backup_path):
            flash('El respaldo no existe', 'error')
            return redirect(url_for('respaldos'))
        
        # Crear respaldo de la base de datos actual antes de restaurar
        crear_respaldo_automatico()
        
        # Restaurar el respaldo
        db_path = 'instance/sistema_ventas.db'
        shutil.copy2(backup_path, db_path)
        
        flash(f'Base de datos restaurada desde {filename}', 'success')
        return redirect(url_for('dashboard'))
        
    except Exception as e:
        flash(f'Error al restaurar respaldo: {str(e)}', 'error')
        return redirect(url_for('respaldos'))

@app.route('/respaldos/descargar/<filename>')
@login_required
def descargar_respaldo(filename):
    """Descargar un respaldo"""
    try:
        from flask import send_file
        
        backup_path = os.path.join('backups', filename)
        if not os.path.exists(backup_path):
            flash('El respaldo no existe', 'error')
            return redirect(url_for('respaldos'))
        
        return send_file(backup_path, as_attachment=True, download_name=filename)
        
    except Exception as e:
        flash(f'Error al descargar respaldo: {str(e)}', 'error')
        return redirect(url_for('respaldos'))

@app.route('/respaldos/eliminar/<filename>')
@login_required
def eliminar_respaldo(filename):
    """Eliminar un respaldo"""
    try:
        backup_path = os.path.join('backups', filename)
        if os.path.exists(backup_path):
            os.remove(backup_path)
            flash(f'Respaldo {filename} eliminado exitosamente', 'success')
        else:
            flash('El respaldo no existe', 'error')
    except Exception as e:
        flash(f'Error al eliminar respaldo: {str(e)}', 'error')
    
    return redirect(url_for('respaldos'))

def crear_usuarios_estaticos():
    """Crear usuarios estáticos si no existen"""
    try:
        if not Usuario.query.filter_by(username='Alonso').first():
            alonso = Usuario(username='Alonso', email='alonso@empresa.com')
            alonso.set_password('123456')
            db.session.add(alonso)
        
        if not Usuario.query.filter_by(username='Andrea').first():
            andrea = Usuario(username='Andrea', email='andrea@empresa.com')
            andrea.set_password('123456')
            db.session.add(andrea)
        
        db.session.commit()
        print("Usuarios estáticos creados: Alonso y Andrea (contraseña: 123456)")
    except Exception as e:
        print(f"Usuarios ya existen o error: {e}")
        db.session.rollback()

def crear_respaldo_automatico():
    """Crear respaldo automático de la base de datos"""
    try:
        from shutil import copy2
        import os
        from datetime import datetime
        
        # Crear directorio de respaldos si no existe
        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        # Nombre del archivo de respaldo con timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f'backup_sistema_ventas_{timestamp}.db'
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # Copiar la base de datos actual
        db_path = 'instance/sistema_ventas.db'
        if os.path.exists(db_path):
            copy2(db_path, backup_path)
            print(f"Respaldo creado: {backup_path}")
            return backup_path
        else:
            print("No se encontró la base de datos para respaldar")
            return None
    except Exception as e:
        print(f"Error al crear respaldo: {e}")
        return None

def es_venta_del_dia_actual(fecha_venta):
    """Verifica si una venta es del día actual"""
    from datetime import date
    hoy = date.today()
    return fecha_venta.date() == hoy

def cargar_datos_existentes():
    """Cargar datos existentes o crear estructura inicial"""
    try:
        # Verificar si la base de datos existe
        db_path = 'instance/sistema_ventas.db'
        if os.path.exists(db_path):
            print("Cargando base de datos existente...")
            # Solo crear las tablas si no existen
            db.create_all()
            print("Base de datos cargada exitosamente")
        else:
            print("Creando nueva base de datos...")
            # Crear directorio instance si no existe
            os.makedirs('instance', exist_ok=True)
            db.create_all()
            print("Base de datos creada exitosamente")
        
        # Crear usuarios estáticos solo si no existen
        crear_usuarios_estaticos()
        
        # Crear respaldo inicial solo si no hay respaldos recientes
        crear_respaldo_si_es_necesario()
        
    except Exception as e:
        print(f"Error al cargar datos: {e}")

def crear_respaldo_si_es_necesario():
    """Crear respaldo solo si es necesario (no hay respaldos recientes)"""
    try:
        import os
        from datetime import datetime, timedelta
        
        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
            # Si no existe el directorio, crear el primer respaldo
            crear_respaldo_automatico()
            return
        
        # Verificar si hay respaldos recientes (últimas 24 horas)
        respaldos_recientes = []
        for filename in os.listdir(backup_dir):
            if filename.endswith('.db'):
                file_path = os.path.join(backup_dir, filename)
                file_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                if file_time > datetime.now() - timedelta(hours=24):
                    respaldos_recientes.append(file_time)
        
        # Si no hay respaldos recientes, crear uno
        if not respaldos_recientes:
            print("No hay respaldos recientes, creando respaldo automático...")
            crear_respaldo_automatico()
        else:
            print(f"Ya existen {len(respaldos_recientes)} respaldos recientes")
            
    except Exception as e:
        print(f"Error al verificar respaldos: {e}")
        # En caso de error, crear respaldo de seguridad
        crear_respaldo_automatico()

if __name__ == '__main__':
    with app.app_context():
        cargar_datos_existentes()
        
    app.run(debug=True)
